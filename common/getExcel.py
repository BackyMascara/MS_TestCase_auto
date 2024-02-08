from openpyxl import load_workbook

def GetExcel(my_time,requirementNames,tableNames,fieldNames):
    """
    用于对单个需求文档，编写Excel测试用例
    :param my_time: 迭代周期时间
    :param requirementNames: 需求数组
    :param tableNames: 表名数组
    :param fieldNames: 表字段数组
    :return:
    """
    # 打开现有的模板Excel文件
    #wb = load_workbook(r'E:\测试用例-fl\爬虫\20240204-20240223\测试用例-自动.xlsx')
    wb = load_workbook(r'../data/测试用例-模板.xlsx')

    # 选择要操作的工作表（默认为第一个）
    ws = wb.active
    i = 2 #测试用例名称的单元格位置
    j = 4 #测试步骤的单元格位置
    TC_name_index = 'A' + str(i)  # 用例名称
    TC_module_index = 'B' + str(i)  # 所属模块
    # 设置要写入的单元格位置及值
    ws[TC_name_index] = '检查“ ' + requirementNames + ' ”需求数据'
    ws[TC_module_index] = '/' + my_time + '/' + requirementNames + ' ( ' + tableNames + ' )'

    # 前两行固定写死
    ws['C2'] = '数据准确性'
    ws['D2'] = '与原网站数据一致'
    ws['C3'] = '与原网站数据一致'
    ws['D3'] = '与原网站数据一致'

    for fieldName in fieldNames:
        TC_step_index = 'C' + str(j)  # 步骤描述
        TC_result_index = 'D' + str(j)  # 预期结果

        #第3行
        ws[TC_step_index] = '检查mysteel_spider.' + tableNames + ' 表 ' + fieldName + ' 字段数据'
        ws[TC_result_index] = '符合前置需求文档字段清洗逻辑和质检逻辑'

        j = j + 1

    # 保存修改并关闭Excel文件
    save_url = r'E:\测试用例-fl\爬虫\{}\测试用例-{}.xlsx'.format(my_time,tableNames)
    wb.save(save_url)
    wb.close()

def GetExcels(my_time,requirementNames,tableNames,fieldNames):
    """
    用于对多个需求文档，编写Excel测试用例
    :param my_time: 迭代周期时间
    :param requirementNames: 需求数组
    :param tableNames: 表名数组
    :param fieldNames: 表字段数组
    :return:
    """
    # 打开现有的模板Excel文件
    #wb = load_workbook(r'E:\测试用例-fl\爬虫\20240204-20240223\测试用例-自动.xlsx')
    wb = load_workbook(r'../data/测试用例-模板.xlsx')

    # 选择要操作的工作表（默认为第一个）
    ws = wb.active
    i = 2  # 测试用例名称的单元格位置
    j = 2  # 测试步骤的单元格位置

    index = 0 # 需求游标
    # 遍历每一个需求
    for requirementName in requirementNames:
        A_index = 'A' + str(i)  # 用例名称
        B_index = 'B' + str(i)  # 所属模块
        C_index = 'C' + str(i)  # 标签
        D_index = 'D' + str(j)  # 步骤描述
        E_index = 'E' + str(j)  # 预期结果
        F_index = 'F' + str(i)  # 编辑模式
        G_index = 'G' + str(i)  # 用例状态
        H_index = 'H' + str(i)  # 责任人
        I_index = 'I' + str(i)  # 用例等级
        J_index = 'J' + str(i)  # 用例分类


        # 设置要写入的单元格位置及值
        ws[A_index] = '检查“ ' + requirementName + ' ”需求数据'
        ws[B_index] = '/' + my_time + '/' + requirementName + ' ( ' + tableNames[index] + ' )'
        ws[C_index] = '数据'
        ws[F_index] = 'STEP'
        ws[G_index] = '未开始'
        ws[H_index] = '樊磊'
        ws[I_index] = 'P0'
        ws[J_index] = '[后端]'

        # 步骤：前两行固定写死（第一行）
        ws[D_index] = '数据准确性'
        ws[E_index] = '与原网站数据一致'

        j = j + 1  # 第二行
        D_index = 'D' + str(j)  # 步骤描述
        E_index = 'E' + str(j)  # 预期结果
        ws[D_index] = '与原网站数据一致'
        ws[E_index] = '与原网站数据一致'

        j = j + 1  # 第三行开始，遍历每个需求的表格字段
        for fieldName in fieldNames[index]:

            D_index = 'D' + str(j)  # 步骤描述
            E_index = 'E' + str(j)  # 预期结果
            ws[D_index] = '检查mysteel_spider.' + tableNames[index] + ' 表 ' + fieldName + ' 字段数据'
            ws[E_index] = '符合前置需求文档字段清洗逻辑和质检逻辑'

            j = j + 1

        index = index + 1
        i = index * 23 + 2  # 下一个合并单元格（下一个需求）
        j = index * 23 + 2

    # 保存修改并关闭Excel文件
    save_url = r'E:\测试用例-fl\爬虫\{}\测试用例-{}.xlsx'.format(my_time, "test")
    wb.save(save_url)
    wb.close()

if __name__ == '__main__':
    my_time = '20240204-20240223'
    fieldNames = [['data_date', 'breed_name', 'maturity_num', 'maturity_num_1m', 'maturity_num_2m', 'remark'],['data_date', 'breed_name', 'maturity_num', 'maturity_num_1m', 'maturity_num_2m', 'remark0207']]
    requirementNames = ['深圳证券交易所','上海证券交易所']
    tableNames = ['eco_spi_power_mddl','eco_spi_capacity_ggdl']
    GetExcels(my_time,requirementNames,tableNames,fieldNames)